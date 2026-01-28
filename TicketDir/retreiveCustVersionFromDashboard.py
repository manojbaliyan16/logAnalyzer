
# import module
from selenium import webdriver
import pandas as pd
import xlrd
import os
import xlsxwriter
import openpyxl
import time
  
# Create the webdriver object. Here the 
# chromedriver is present in the driver 
# folder of the root directory.
#_SW_ID_="AI_PRJ_RN_AIVI_19.3V22"

def getBranch(_SW_ID_):   
    # initializing substrings
    res="UnknownBoschVersion"
    try:
        sub1 = "AI_PRJ_RN_AIVI_"
        sub2 = "V"
          
        # getting index of substrings
        idx1 = _SW_ID_.index(sub1)
        idx2 = _SW_ID_.find(sub2, _SW_ID_.find(sub2)+1)
        print(idx1)
        print(idx2)
          
        # length of substring 1 is added to
        # get string from next character
        res = _SW_ID_[idx1 + len(sub1): idx2]
          
        # printing result
        print("The extracted branch : " + res)
    except:
        print("Unknown Bosch Version")
    return res
    
def copyAsExcel(branch_link,_SW_ID_):   
    # Assign the table data to a Pandas dataframe
    table = pd.read_html(branch_link)[0]
  
    # Print the dataframe
    print(table)

    # Store the dataframe in Excel file
    
    table.to_excel("CustomerVersion_Information.xlsx")
    return
    
def getcuVersion(_SW_ID_):
    custVersion="Unavailable"
    
    import openpyxl

    book = openpyxl.load_workbook('CustomerVersion_Information.xlsx')

    sheet = book.active

    a1 = sheet['A1']
    a2 = sheet['A2']
    a3 = sheet.cell(row=3, column=1)

    print(a1.value)
    print(a2.value) 
    print(a3.value)
    if(sheet["B1"].value=="Release Label" and sheet["D1"].value=="Overall Version"):
        buildVersionColumn="B"
        customerVersionColumn="D"
    elif(sheet["C1"].value=="Release Label" and sheet["E1"].value=="Overall Version"):
        buildVersionColumn="C"
        customerVersionColumn="E"
    count=1    
    print(sheet.max_row)
    for r in range(0,sheet.max_row):
        if(count==1):
            count=count+1
            continue
        col_name=buildVersionColumn+str(count)
        SW_ID_in_sheet=sheet[col_name].value
        if(_SW_ID_==str(SW_ID_in_sheet)):
            col_name=customerVersionColumn+str(count)
            custVersion=sheet[col_name].value
            break
        count=count+1 
    return custVersion
    
def getIntegrationDashboardWebPage(_SW_ID_,branch_type):
    Integration_DashBoard_link="http://10.165.234.103:9090/IntegrationDashBoard/"
    driver=webdriver.Firefox(executable_path="geckodriver.exe")
    driver.implicitly_wait(0.5) 
    driver.get(Integration_DashBoard_link)

    branch=""
    # Maximize the window and let code stall 
    # for 10s to properly maximise the window.
    driver.maximize_window()
    #AI_PRJ_RN_AIVI_19.3V22
    branch=getBranch(_SW_ID_)
    if(branch=="UnknownBoschVersion"):
        return branch
    #branch="18.1"
    if(branch_type=="cc"):
        try:
            button_text="//button[text()='ai_prj_rn_aivi_"+branch+"_int']"
            print(button_text)
            l=driver.find_element_by_xpath(button_text)
        except:
            button_text="//div[@class='GitBr']/button[text()='rn_aivi_"+branch+"_stabi']"
            print(button_text)
            l=driver.find_element_by_xpath(button_text)
    else:
        button_text="//div[@class='GitBr']/button[text()='rn_aivi_"+branch+"_stabi']"
        print(button_text)
        l=driver.find_element_by_xpath(button_text)        
    l.click()
    tab1 = driver.window_handles[0]
    tab2 = driver.window_handles[1]

    driver.switch_to.window(tab2) # switch to new tab

    # do your stuff here
    while(Integration_DashBoard_link==driver.current_url):
            print("")
    branch_link=driver.current_url
    print(branch_link)
    driver.close()
    driver.switch_to.window(tab1) # switch to original tab
    driver.close() # close tab
    return branch_link
    
def getCustomerVersionFromDashboard(_SW_ID_,branch_type):
    branch_link=getIntegrationDashboardWebPage(_SW_ID_,branch_type)
    if(branch_link=="UnknownBoschVersion"):
        print("---")
        return branch_link
    print("branch_link: "+branch_link)
    copyAsExcel(branch_link,_SW_ID_)
    customer_version=getcuVersion(_SW_ID_)
    print("customer_version in rereiving file")
    print(customer_version)
    return customer_version
#getCustomerVersionFromDashboard(_SW_ID_)